from django.shortcuts import render
from django.views.generic import View, TemplateView, CreateView, ListView, DeleteView, UpdateView, FormView, RedirectView
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseRedirect, HttpResponse
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import *
import openpyxl
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator


class IndexView(TemplateView):
    template_name = "index.html"


@method_decorator(csrf_exempt, name='dispatch')
class LoginView(TemplateView):

    @staticmethod
    def get(request):
        if request.user.is_authenticated:
            return redirect('index')
        else:
            return render(request, 'login.html')

    @staticmethod
    def post(request):
        form_data = request.POST
        username = form_data.get('username', None)
        password = form_data.get('password', None)
        next_url = form_data.get(
            'next', None) if form_data.get('next') else '/'
        if not username and password:
            return render(request, 'login.html', {'message': "* Invalid User"})
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('index')
        else:
            return render(request, 'login.html', {'message': "* Invalid Username or password"})


class LogoutView(RedirectView):
    url = reverse_lazy('login')

    def get(self, request, *args, **kwargs):
        logout(request)
        return redirect('index')


@method_decorator(csrf_exempt, name='dispatch')
class StreamerCreateView(TemplateView):
    template_name = 'streamers/create.html'

    @staticmethod
    def post(request):
        form_data = request.POST
        form_files = request.FILES
        streamer = Streamer.objects.create(
            name=form_data.get('name', "None"),
            details=form_data.get('detail'),
            url=form_data.get('url'),
            movies=form_files.get('files')
        )
        streamer.save()

        path = streamer.movies.path
        try:
            wb_obj = openpyxl.load_workbook(path)
            sheet_obj = wb_obj.active
            m_row = sheet_obj.max_row

            for i in range(1, m_row + 1):
                cell_obj = sheet_obj.cell(row=i, column=1)
                movieName = MovieName.objects.create(
                    name=cell_obj.value,
                    streamer=streamer
                )
            movieName.save()
            return JsonResponse({
                'status': True,
                'message': "Streamer Created!!!"
            })
        except:
            return JsonResponse({
                'status': False,
                'message': "You are not allowed to do this!!!"
            })


@method_decorator(csrf_exempt, name='dispatch')
class StreamerListView(ListView):

    model = Streamer
    template_name = 'streamers/list.html'
    context_object_name = 'streamers'
    paginate_by = 5

    def get_context_data(self, **kwargs):
        context = super(StreamerListView, self).get_context_data(**kwargs)
        streamers = self.get_queryset()
        page = self.request.GET.get('page')
        paginator = Paginator(streamers, self.paginate_by)
        try:
            streamers = paginator.page(page)
        except PageNotAnInteger:
            streamers = paginator.page(1)
        except EmptyPage:
            streamers = paginator.page(paginator.num_pages)
        context['streamers'] = streamers
        return context


@method_decorator(csrf_exempt, name='dispatch')
class StreamerUpdateView(UpdateView):

    model = Streamer
    template_name = 'streamers/update.html'
    context_object_name = 'streamer'
    fields = ('name', 'movies')

    def get_success_url(self):
        return reverse_lazy('list-streamer')


@method_decorator(csrf_exempt, name='dispatch')
class StreamerDeleteView(DeleteView):
    model = Streamer
    template_name = 'streamers/delete.html'
    success_url = reverse_lazy('list-streamer')


@method_decorator(csrf_exempt, name='dispatch')
class MovieDetailView(TemplateView):
    template_name = 'movie-details'

    @staticmethod
    def post(request):
        form_data = request.POST
        s_query = form_data.get('name')
        if (form_data.get('starting') == 'on'):
            list_of_movies = MovieName.objects.filter(
                name__istartswith=s_query)
        else:
            list_of_movies = MovieName.objects.filter(name__icontains=s_query)
        return render(request, 'movie-details.html', {'movies': list_of_movies, 'keyword': s_query})
